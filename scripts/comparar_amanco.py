"""
Comparador de Preços Amanco — Versão adaptada para Streamlit.
Compara preços entre um arquivo Excel e um PDF de fornecedor.
"""

import re
import io
import pandas as pd
import pdfplumber
import warnings
from openpyxl.styles import PatternFill, Font

warnings.filterwarnings("ignore")


class PriceComparator:
    """Compara preços entre arquivo PDF e planilha Excel."""

    def __init__(self, excel_file, pdf_file):
        """
        Inicializa o comparador com file-like objects (uploads do Streamlit).

        Args:
            excel_file: UploadedFile do Excel
            pdf_file:   UploadedFile do PDF
        """
        self.excel_file = excel_file
        self.pdf_file = pdf_file
        self.df_excel = None
        self.df_pdf = None
        self.logs = []  # acumula mensagens de log

    # ----- utilidades -----
    def _log(self, msg: str):
        self.logs.append(msg)

    @staticmethod
    def extract_product_code(product_text):
        if pd.isna(product_text):
            return None
        matches = re.findall(r"\b(\d{4,6})\b", str(product_text))
        return matches[-1] if matches else None

    @staticmethod
    def clean_price(price_text):
        if pd.isna(price_text):
            return None
        price_str = (
            str(price_text)
            .replace("R$", "")
            .replace(".", "")
            .replace(",", ".")
            .strip()
        )
        try:
            return float(price_str)
        except ValueError:
            return None

    # ----- leitura -----
    def read_excel(self):
        self._log("📊 Lendo arquivo Excel…")
        self.df_excel = pd.read_excel(self.excel_file)
        self.df_excel["Código Extraído"] = self.df_excel["Produto"].apply(
            self.extract_product_code
        )

        if "Compra" in self.df_excel.columns:
            self.df_excel["Compra"] = pd.to_numeric(
                self.df_excel["Compra"], errors="coerce"
            )
        if "Preço" in self.df_excel.columns:
            self.df_excel["Preço"] = pd.to_numeric(
                self.df_excel["Preço"], errors="coerce"
            )
        elif "Valor de compra" in self.df_excel.columns:
            self.df_excel["Valor de compra"] = pd.to_numeric(
                self.df_excel["Valor de compra"], errors="coerce"
            )

        self._log(f"✅ Excel: {len(self.df_excel)} linhas, "
                   f"{self.df_excel['Código Extraído'].notna().sum()} códigos extraídos")

    def read_pdf(self):
        self._log("📄 Lendo arquivo PDF…")
        pdf_data = []

        with pdfplumber.open(self.pdf_file) as pdf:
            self._log(f"   Páginas: {len(pdf.pages)}")

            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()

                if tables:
                    for table in tables:
                        if not table or len(table) < 2:
                            continue

                        header = [str(h).lower() if h else "" for h in table[0]]
                        codigo_idx = descricao_idx = qtde_idx = preco_idx = None

                        for i, h in enumerate(header):
                            if "código" in h or "codigo" in h:
                                codigo_idx = i
                            if "descrição" in h or "descricao" in h or "produto" in h:
                                descricao_idx = i
                            if "qtde" in h or "quantidade" in h:
                                qtde_idx = i
                            if "preço líq" in h or "preco liq" in h or "preço liq" in h:
                                preco_idx = i

                        if None in (codigo_idx, qtde_idx, preco_idx):
                            continue

                        for row in table[1:]:
                            try:
                                if len(row) > max(codigo_idx, qtde_idx, preco_idx):
                                    codigo = str(row[codigo_idx]).strip() if row[codigo_idx] else None
                                    descricao = (
                                        str(row[descricao_idx]).strip()
                                        if descricao_idx is not None and row[descricao_idx]
                                        else ""
                                    )
                                    qtde = row[qtde_idx]
                                    preco = row[preco_idx]

                                    if codigo and re.match(r"^\d{4,6}$", codigo):
                                        pdf_data.append({
                                            "Código_PDF": codigo,
                                            "Descricao_PDF": descricao,
                                            "Qtde_PDF": qtde,
                                            "Preço_Líq_PDF": preco,
                                        })
                            except Exception:
                                continue

                if not tables or not pdf_data:
                    text = page.extract_text()
                    if text:
                        for line in text.split("\n"):
                            match = re.match(
                                r"^(\d{4,6})\s+(.*?)(?:BR\d+|0\d+)\s*-\s*.*?\s+(\d+)\s+R\$\s*([\d.,]+)",
                                line,
                            )
                            if match:
                                pdf_data.append({
                                    "Código_PDF": match.group(1),
                                    "Descricao_PDF": match.group(2).strip(),
                                    "Qtde_PDF": match.group(3),
                                    "Preço_Líq_PDF": match.group(4),
                                })

        self.df_pdf = pd.DataFrame(pdf_data)

        if not self.df_pdf.empty:
            self.df_pdf["Qtde_PDF"] = pd.to_numeric(self.df_pdf["Qtde_PDF"], errors="coerce")
            self.df_pdf["Preço_Líq_PDF"] = self.df_pdf["Preço_Líq_PDF"].apply(self.clean_price)

            def adjust_eletroduto(row):
                desc = str(row.get("Descricao_PDF", "")).upper()
                if "ELETRODUTO" in desc:
                    m = re.search(r"X(\d+)M\b", desc)
                    if m:
                        fator = float(m.group(1))
                        if pd.notna(row["Qtde_PDF"]):
                            row["Qtde_PDF"] = row["Qtde_PDF"] / fator
                        if pd.notna(row["Preço_Líq_PDF"]):
                            row["Preço_Líq_PDF"] = row["Preço_Líq_PDF"] * fator
                return row

            self.df_pdf = self.df_pdf.apply(adjust_eletroduto, axis=1)
            self.df_pdf = self.df_pdf.drop("Descricao_PDF", axis=1, errors="ignore")
            self.df_pdf = self.df_pdf.drop_duplicates(subset=["Código_PDF"], keep="first")

        self._log(f"✅ PDF: {len(self.df_pdf)} itens extraídos")

    # ----- processamento -----
    def merge_data(self):
        self._log("🔄 Combinando dados…")

        if self.df_pdf.empty:
            df_result = self.df_excel.copy()
            df_result["Qtde_PDF"] = None
            df_result["Preço_Líq_PDF"] = None
        else:
            df_result = self.df_excel.merge(
                self.df_pdf,
                left_on="Código Extraído",
                right_on="Código_PDF",
                how="left",
            )
            if "Código_PDF" in df_result.columns:
                df_result = df_result.drop("Código_PDF", axis=1)

        matches = df_result["Qtde_PDF"].notna().sum()
        self._log(f"✅ {matches} correspondências encontradas")
        return df_result

    def calculate_differences(self, df_result):
        self._log("🔢 Calculando diferenças…")

        if "Compra" in df_result.columns:
            df_result["Diferença de Qtde"] = df_result["Qtde_PDF"] - df_result["Compra"]
        else:
            df_result["Diferença de Qtde"] = None

        preco_col = "Preço" if "Preço" in df_result.columns else "Valor de compra"
        if preco_col in df_result.columns:
            df_result["Diferença de Preço"] = df_result["Preço_Líq_PDF"] - df_result[preco_col]
        else:
            df_result["Diferença de Preço"] = None

        return df_result

    # ----- saída Excel -----
    def generate_excel_bytes(self, df_result) -> bytes:
        """Gera o Excel com formatação e retorna como bytes."""
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_result.to_excel(writer, index=False, sheet_name="Comparação")
            ws = writer.sheets["Comparação"]

            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            bold_font = Font(bold=True)

            for cell in ws[1]:
                cell.font = bold_font

            headers = [cell.value for cell in ws[1]]

            try:
                qtde_col = headers.index("Diferença de Qtde") + 1
                preco_col = headers.index("Diferença de Preço") + 1

                for row in range(2, len(df_result) + 2):
                    for col_idx, fills in [(qtde_col, (green_fill, red_fill)),
                                           (preco_col, (red_fill, green_fill))]:
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            try:
                                val = float(cell.value)
                                if val > 0:
                                    cell.fill = fills[0]
                                elif val < 0:
                                    cell.fill = fills[1]
                            except (ValueError, TypeError):
                                pass
            except ValueError:
                pass

            for column in ws.columns:
                max_len = max(
                    (len(str(c.value)) for c in column if c.value is not None),
                    default=8,
                )
                ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)

        return output.getvalue()

    # ----- execução completa -----
    def compare(self):
        """
        Executa a comparação completa.

        Returns:
            tuple: (df_result, excel_bytes, logs)
        """
        self._log("🚀 Iniciando comparação de preços…")
        self.read_excel()
        self.read_pdf()
        df_result = self.merge_data()
        df_result = self.calculate_differences(df_result)
        excel_bytes = self.generate_excel_bytes(df_result)
        self._log("✨ Comparação concluída com sucesso!")
        return df_result, excel_bytes, self.logs


def run(inputs: dict):
    """
    Ponto de entrada chamado pelo app.py.

    Args:
        inputs: dict com chaves 'excel_file' e 'pdf_file' (UploadedFile)

    Returns:
        dict com chaves: df_result, excel_bytes, logs, summary
    """
    comparator = PriceComparator(inputs["excel_file"], inputs["pdf_file"])
    df_result, excel_bytes, logs = comparator.compare()

    matches = df_result["Qtde_PDF"].notna().sum()
    total = len(df_result)
    qtde_dif = (df_result.get("Diferença de Qtde", pd.Series(dtype=float)) != 0).sum()
    preco_dif = (df_result.get("Diferença de Preço", pd.Series(dtype=float)) != 0).sum()

    summary = {
        "total": total,
        "matches": int(matches),
        "no_match": int(total - matches),
        "qtde_divergente": int(qtde_dif),
        "preco_divergente": int(preco_dif),
    }

    return {
        "df_result": df_result,
        "excel_bytes": excel_bytes,
        "logs": logs,
        "summary": summary,
    }
