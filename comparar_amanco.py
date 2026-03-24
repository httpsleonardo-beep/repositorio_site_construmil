import tkinter as tk
from tkinter import filedialog, messagebox
import re
import pandas as pd
import pdfplumber
from pathlib import Path
import warnings
from openpyxl.styles import PatternFill, Font

warnings.filterwarnings('ignore')


class PriceComparator:
    """
    Classe para comparar preços entre arquivo PDF e planilha Excel.
    """

    def __init__(self, excel_path, pdf_path):
        """
        Inicializa o comparador com os caminhos dos arquivos.

        Args:
            excel_path (str): Caminho para o arquivo Excel
            pdf_path (str): Caminho para o arquivo PDF
        """
        self.excel_path = excel_path
        self.pdf_path = pdf_path
        self.df_excel = None
        self.df_pdf = None

    def extract_product_code(self, product_text):
        """
        Extrai o código do produto do texto da coluna Produto.
        Procura por sequências de 4-6 dígitos no final ou meio do texto.

        Args:
            product_text (str): Texto contendo o código do produto

        Returns:
            str: Código do produto encontrado ou None
        """
        if pd.isna(product_text):
            return None

        # Procura por sequências de 4-6 dígitos
        matches = re.findall(r'\b(\d{4,6})\b', str(product_text))
        # Retorna o último código encontrado (geralmente está no final)
        return matches[-1] if matches else None

    def clean_price(self, price_text):
        """
        Limpa e converte texto de preço para float.
        Remove 'R$', pontos de milhar e substitui vírgula por ponto.

        Args:
            price_text (str): Texto do preço (ex: 'R$ 5,01')

        Returns:
            float: Valor numérico do preço
        """
        if pd.isna(price_text):
            return None

        # Remove R$, espaços e converte vírgula para ponto
        price_str = str(price_text).replace('R$', '').replace('.', '').replace(',', '.').strip()

        try:
            return float(price_str)
        except ValueError:
            return None

    def read_excel(self):
        """
        Lê a planilha Excel e extrai os códigos dos produtos.
        """
        print(f"📊 Lendo arquivo Excel: {self.excel_path}")

        try:
            self.df_excel = pd.read_excel(self.excel_path)

            print(f"   Colunas encontradas: {list(self.df_excel.columns)}")

            # Extrai código do produto da coluna 'Produto'
            self.df_excel['Código Extraído'] = self.df_excel['Produto'].apply(
                self.extract_product_code
            )

            # Limpa a coluna 'Compra' se existir
            if 'Compra' in self.df_excel.columns:
                self.df_excel['Compra'] = pd.to_numeric(
                    self.df_excel['Compra'], errors='coerce'
                )

            # Limpa a coluna 'Preço' ou 'Valor de compra' se existir
            if 'Preço' in self.df_excel.columns:
                self.df_excel['Preço'] = pd.to_numeric(
                    self.df_excel['Preço'], errors='coerce'
                )
            elif 'Valor de compra' in self.df_excel.columns:
                self.df_excel['Valor de compra'] = pd.to_numeric(
                    self.df_excel['Valor de compra'], errors='coerce'
                )

            print(f"✅ Excel lido com sucesso: {len(self.df_excel)} linhas")
            print(f"   Códigos extraídos: {self.df_excel['Código Extraído'].notna().sum()}")

            # Mostra alguns exemplos
            if not self.df_excel.empty:
                print(f"\n   Exemplos de códigos extraídos:")
                for idx, row in self.df_excel.head(3).iterrows():
                    print(f"   - {row.get('Produto', 'N/A')} → Código: {row['Código Extraído']}")

        except Exception as e:
            print(f"❌ Erro ao ler Excel: {e}")
            raise

    def read_pdf(self):
        """
        Lê o arquivo PDF e extrai as informações de código, quantidade e preço.
        Usa múltiplas estratégias para extrair os dados.
        """
        print(f"\n📄 Lendo arquivo PDF: {self.pdf_path}")

        try:
            pdf_data = []

            with pdfplumber.open(self.pdf_path) as pdf:
                print(f"   Total de páginas: {len(pdf.pages)}")

                for page_num, page in enumerate(pdf.pages, 1):
                    print(f"   Processando página {page_num}...")

                    # Estratégia 1: Extrai tabelas estruturadas
                    tables = page.extract_tables()

                    if tables:
                        for table_idx, table in enumerate(tables):
                            if not table or len(table) < 2:
                                continue

                            print(f"      Tabela {table_idx + 1} encontrada com {len(table)} linhas")

                            # Mostra o cabeçalho para debug
                            header = [str(h).lower() if h else '' for h in table[0]]
                            print(f"      Cabeçalho: {header}")

                            # Procura pelos índices das colunas
                            codigo_idx = None
                            descricao_idx = None
                            qtde_idx = None
                            preco_idx = None

                            for i, h in enumerate(header):
                                if 'código' in h or 'codigo' in h:
                                    codigo_idx = i
                                if 'descrição' in h or 'descricao' in h or 'produto' in h:
                                    descricao_idx = i
                                if 'qtde' in h or 'quantidade' in h:
                                    qtde_idx = i
                                if 'preço líq' in h or 'preco liq' in h or 'preço liq' in h:
                                    preco_idx = i

                            print(
                                f"      Índices: Código={codigo_idx}, Descrição={descricao_idx}, Qtde={qtde_idx}, Preço={preco_idx}")

                            if None in (codigo_idx, qtde_idx, preco_idx):
                                print(f"      ⚠️  Colunas necessárias não encontradas nesta tabela")
                                continue

                            # Processa cada linha da tabela
                            for row_idx, row in enumerate(table[1:], 1):
                                try:
                                    if len(row) > max(codigo_idx, qtde_idx, preco_idx):
                                        codigo = str(row[codigo_idx]).strip() if row[codigo_idx] else None
                                        descricao = str(row[descricao_idx]).strip() if descricao_idx is not None and \
                                                                                       row[descricao_idx] else ""
                                        qtde = row[qtde_idx]
                                        preco = row[preco_idx]

                                        # Valida se há código (4-6 dígitos)
                                        if codigo and re.match(r'^\d{4,6}$', codigo):
                                            pdf_data.append({
                                                'Código_PDF': codigo,
                                                'Descricao_PDF': descricao,
                                                'Qtde_PDF': qtde,
                                                'Preço_Líq_PDF': preco
                                            })

                                except Exception as e:
                                    continue

                    # Estratégia 2: Se não encontrou tabelas, tenta extrair texto
                    if not tables or not pdf_data:
                        text = page.extract_text()
                        if text:
                            print(f"      Tentando extrair do texto bruto...")
                            # Procura por padrões de linhas de produtos
                            lines = text.split('\n')
                            for line in lines:
                                # Regex modificada para tentar pegar a descrição também no Grupo 2
                                match = re.match(
                                    r'^(\d{4,6})\s+(.*?)(?:BR\d+|0\d+)\s*-\s*.*?\s+(\d+)\s+R\$\s*([\d.,]+)',
                                    line
                                )
                                if match:
                                    codigo = match.group(1)
                                    descricao = match.group(2).strip()
                                    qtde = match.group(3)
                                    preco = match.group(4)

                                    pdf_data.append({
                                        'Código_PDF': codigo,
                                        'Descricao_PDF': descricao,
                                        'Qtde_PDF': qtde,
                                        'Preço_Líq_PDF': preco
                                    })

            self.df_pdf = pd.DataFrame(pdf_data)

            # Limpa os valores de quantidade e preço
            if not self.df_pdf.empty:
                self.df_pdf['Qtde_PDF'] = pd.to_numeric(
                    self.df_pdf['Qtde_PDF'], errors='coerce'
                )
                self.df_pdf['Preço_Líq_PDF'] = self.df_pdf['Preço_Líq_PDF'].apply(
                    self.clean_price
                )

                # ==========================================
                # NOVO: LÓGICA DE CONVERSÃO DO ELETRODUTO
                # ==========================================
                def adjust_eletroduto(row):
                    desc = str(row.get('Descricao_PDF', '')).upper()
                    if 'ELETRODUTO' in desc:
                        # Busca por um padrão como "X25M" ou "X50M" na descrição
                        match = re.search(r'X(\d+)M\b', desc)
                        if match:
                            fator = float(match.group(1))
                            if pd.notna(row['Qtde_PDF']):
                                row['Qtde_PDF'] = row['Qtde_PDF'] / fator
                            if pd.notna(row['Preço_Líq_PDF']):
                                row['Preço_Líq_PDF'] = row['Preço_Líq_PDF'] * fator
                    return row

                # Aplica a função para corrigir os rolos
                self.df_pdf = self.df_pdf.apply(adjust_eletroduto, axis=1)

                # Remove a coluna Descricao_PDF que era só um auxiliar
                self.df_pdf = self.df_pdf.drop('Descricao_PDF', axis=1, errors='ignore')

                # Remove duplicatas mantendo a primeira ocorrência
                self.df_pdf = self.df_pdf.drop_duplicates(subset=['Código_PDF'], keep='first')

            print(f"✅ PDF lido com sucesso: {len(self.df_pdf)} itens únicos extraídos")

            # Mostra alguns exemplos
            if not self.df_pdf.empty:
                print(f"\n   Exemplos de itens extraídos do PDF:")
                for idx, row in self.df_pdf.head(3).iterrows():
                    print(
                        f"   - Código: {row['Código_PDF']} | Qtde: {row['Qtde_PDF']} | Preço: R$ {row['Preço_Líq_PDF']}")
            else:
                print("\n   ⚠️  ATENÇÃO: Nenhum item foi extraído do PDF!")
                print("   Verifique se o PDF contém tabelas ou se a estrutura está correta.")

        except Exception as e:
            print(f"❌ Erro ao ler PDF: {e}")
            raise

    def merge_data(self):
        """
        Combina os dados do Excel e PDF baseado no código do produto.

        Returns:
            pd.DataFrame: DataFrame combinado com todas as informações
        """
        print("\n🔄 Combinando dados...")

        if self.df_pdf.empty:
            print("   ⚠️  PDF vazio - adicionando colunas vazias")
            # Se o PDF está vazio, apenas adiciona colunas vazias
            df_result = self.df_excel.copy()
            df_result['Qtde_PDF'] = None
            df_result['Preço_Líq_PDF'] = None
        else:
            # Faz o merge usando o código extraído do Excel e o código do PDF
            df_result = self.df_excel.merge(
                self.df_pdf,
                left_on='Código Extraído',
                right_on='Código_PDF',
                how='left'
            )

            # Remove a coluna auxiliar Código_PDF
            if 'Código_PDF' in df_result.columns:
                df_result = df_result.drop('Código_PDF', axis=1)

            matches = df_result['Qtde_PDF'].notna().sum()
            print(f"✅ Combinação concluída: {matches} correspondências encontradas")

            if matches == 0:
                print("   ⚠️  Nenhuma correspondência encontrada!")
                print("   Verifique se os códigos do Excel e PDF estão no mesmo formato.")

        return df_result

    def calculate_differences(self, df_result):
        """
        Calcula as diferenças entre Excel e PDF.

        Args:
            df_result (pd.DataFrame): DataFrame com os dados combinados

        Returns:
            pd.DataFrame: DataFrame com as colunas de diferença adicionadas
        """
        print("\n🔢 Calculando diferenças...")

        # Diferença de Quantidade: Qtde_PDF - Compra
        if 'Compra' in df_result.columns:
            df_result['Diferença de Qtde'] = df_result['Qtde_PDF'] - df_result['Compra']
        else:
            df_result['Diferença de Qtde'] = None

        # Diferença de Preço: Preço_Líq_PDF - Preço (ou Valor de compra)
        preco_col = 'Preço' if 'Preço' in df_result.columns else 'Valor de compra'

        if preco_col in df_result.columns:
            df_result['Diferença de Preço'] = df_result['Preço_Líq_PDF'] - df_result[preco_col]
        else:
            df_result['Diferença de Preço'] = None
            print("   ⚠️  Coluna de 'Preço' ou 'Valor de compra' não encontrada no Excel.")

        # Estatísticas das diferenças
        qtde_dif = df_result['Diferença de Qtde'].notna().sum()
        preco_dif = df_result['Diferença de Preço'].notna().sum()

        # Diferenças significativas (não zero)
        qtde_com_dif = (df_result['Diferença de Qtde'] != 0).sum()
        preco_com_dif = (df_result['Diferença de Preço'] != 0).sum()

        print(f"   Diferenças de Quantidade calculadas: {qtde_dif}")
        print(f"   Com diferença não-zero: {qtde_com_dif}")
        print(f"   Diferenças de Preço calculadas: {preco_dif}")
        print(f"   Com diferença não-zero: {preco_com_dif}")

        return df_result

    def save_result(self, df_result, output_path=None):
        """
        Salva o resultado em um novo arquivo Excel com formatação condicional.

        Args:
            df_result (pd.DataFrame): DataFrame com os dados combinados
            output_path (str): Caminho do arquivo de saída (opcional)
        """
        if output_path is None:
            # Gera nome automático baseado no arquivo Excel original
            base_path = Path(self.excel_path)
            output_path = base_path.parent / f"{base_path.stem}_comparacao.xlsx"

        print(f"\n💾 Salvando resultado em: {output_path}")

        try:
            # Cria o writer do Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_result.to_excel(writer, index=False, sheet_name='Comparação')

                # Ajusta largura das colunas
                worksheet = writer.sheets['Comparação']

                # Formatações
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                bold_font = Font(bold=True)

                # Aplica negrito no cabeçalho
                for cell in worksheet[1]:
                    cell.font = bold_font

                # Encontra as colunas de diferença
                headers = [cell.value for cell in worksheet[1]]

                try:
                    qtde_dif_col = headers.index('Diferença de Qtde') + 1
                    preco_dif_col = headers.index('Diferença de Preço') + 1

                    # Aplica formatação condicional nas colunas de diferença
                    for row in range(2, len(df_result) + 2):
                        # Diferença de Quantidade
                        qtde_cell = worksheet.cell(row=row, column=qtde_dif_col)
                        if qtde_cell.value is not None:
                            try:
                                val = float(qtde_cell.value)
                                if val > 0:
                                    qtde_cell.fill = green_fill  # Mais no PDF
                                elif val < 0:
                                    qtde_cell.fill = red_fill  # Menos no PDF
                            except (ValueError, TypeError):
                                pass

                        # Diferença de Preço
                        preco_cell = worksheet.cell(row=row, column=preco_dif_col)
                        if preco_cell.value is not None:
                            try:
                                val = float(preco_cell.value)
                                if val > 0:
                                    preco_cell.fill = red_fill  # Preço maior no PDF
                                elif val < 0:
                                    preco_cell.fill = green_fill  # Preço menor no PDF
                            except (ValueError, TypeError):
                                pass

                except ValueError:
                    print("   ⚠️  Colunas de diferença não encontradas para formatação")

                # Ajusta largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"✅ Arquivo salvo com sucesso!")
            print(f"\n📊 Resumo:")
            print(f"   Total de produtos: {len(df_result)}")
            print(f"   Com correspondência no PDF: {df_result['Qtde_PDF'].notna().sum()}")
            print(f"   Sem correspondência: {df_result['Qtde_PDF'].isna().sum()}")

            # Resumo das diferenças
            qtde_maior = (df_result['Diferença de Qtde'] > 0).sum()
            qtde_menor = (df_result['Diferença de Qtde'] < 0).sum()
            preco_maior = (df_result['Diferença de Preço'] > 0).sum()
            preco_menor = (df_result['Diferença de Preço'] < 0).sum()

            print(f"\n📈 Diferenças:")
            print(f"   Quantidade maior no PDF: {qtde_maior}")
            print(f"   Quantidade menor no PDF: {qtde_menor}")
            print(f"   Preço maior no PDF: {preco_maior}")
            print(f"   Preço menor no PDF: {preco_menor}")

        except Exception as e:
            print(f"❌ Erro ao salvar arquivo: {e}")
            raise

    def compare(self, output_path=None):
        """
        Executa o processo completo de comparação.

        Args:
            output_path (str): Caminho do arquivo de saída (opcional)
        """
        print("🚀 Iniciando comparação de preços...\n")

        # Executa todas as etapas
        self.read_excel()
        self.read_pdf()
        df_result = self.merge_data()
        df_result = self.calculate_differences(df_result)
        self.save_result(df_result, output_path)

        print("\n✨ Processo concluído com sucesso!")


def main():
    """
    Função principal para executar o script com Tkinter.
    """
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal

    # Seleciona o arquivo Excel
    excel_path = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
    )
    if not excel_path:
        messagebox.showinfo("Cancelado", "Operação cancelada.")
        return

    # Seleciona o arquivo PDF
    pdf_path = filedialog.askopenfilename(
        title="Selecione o arquivo PDF",
        filetypes=[("Arquivos PDF", "*.pdf")]
    )
    if not pdf_path:
        messagebox.showinfo("Cancelado", "Operação cancelada.")
        return

    # Seleciona o caminho de saída
    output_path = filedialog.asksaveasfilename(
        title="Salvar comparação como",
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )
    if not output_path:
        messagebox.showinfo("Cancelado", "Operação cancelada.")
        return

    try:
        comparator = PriceComparator(excel_path, pdf_path)
        comparator.compare(output_path)
        messagebox.showinfo("Sucesso", f"Comparação concluída e salva em:\n{output_path}")

    except FileNotFoundError as e:
        messagebox.showerror("Erro", f"Arquivo não encontrado: {e}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro durante a execução: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()